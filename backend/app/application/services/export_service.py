"""Export service for search results"""
from typing import List, Optional, Dict, Any
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import csv
from datetime import datetime
from app.domain.entities.search import Search, Resume
from app.domain.entities.candidate import Candidate
from app.core.logging import get_logger
from app.core.exceptions import NotFoundException

logger = get_logger(__name__)

# PDF generation (optional)
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab not available, PDF export disabled")


class ExportService:
    """Service for exporting search results"""
    
    def _validate_search_id(self, search_id: str) -> None:
        """Validate search_id format (ObjectId)"""
        from bson import ObjectId
        try:
            ObjectId(search_id)
        except Exception:
            raise NotFoundException(f"Invalid search ID format: {search_id}")
    
    async def export_to_excel(self, search_id: str) -> BytesIO:
        """Export search results to Excel"""
        # Validate search_id format
        self._validate_search_id(search_id)
        
        # Get search
        search = await Search.get(search_id)
        if not search:
            raise NotFoundException("Search not found")
        
        # Get all resumes
        resumes = await Resume.find({"search_id": str(search.id)}).sort(-Resume.ai_score).to_list()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumes"
        
        # Enhanced headers with detailed evaluation
        headers = [
            "ID", "HH ID", "Name", "Age", "City", "Title", "Salary", "Currency",
            "Preliminary Score", "AI Score", "Match %", "Status", "Tags",
            "AI Summary", "Match Explanation", "Strengths", "Weaknesses", "Recommendation",
            "Technical Skills Score", "Experience Score", "Education Score", "Soft Skills Score",
            "Red Flags", "AI Questions", "AI Generated", "Analyzed", "Created At"
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
        
        # Add data with detailed evaluation
        for resume in resumes:
            # Get candidate info if exists
            candidate = await Candidate.find_one({"resume_id": str(resume.id)})
            status = candidate.status if candidate else "new"
            tags = ", ".join(candidate.tags) if candidate and candidate.tags else ""
            
            # Extract evaluation details
            tech_score = ""
            exp_score = ""
            edu_score = ""
            soft_score = ""
            
            if resume.evaluation_details:
                if "technical_skills" in resume.evaluation_details:
                    tech_data = resume.evaluation_details["technical_skills"]
                    if isinstance(tech_data, dict) and "score" in tech_data:
                        tech_score = tech_data["score"]
                
                if "experience" in resume.evaluation_details:
                    exp_data = resume.evaluation_details["experience"]
                    if isinstance(exp_data, dict) and "score" in exp_data:
                        exp_score = exp_data["score"]
                
                if "education" in resume.evaluation_details:
                    edu_data = resume.evaluation_details["education"]
                    if isinstance(edu_data, dict) and "score" in edu_data:
                        edu_score = edu_data["score"]
                
                if "soft_skills" in resume.evaluation_details:
                    soft_data = resume.evaluation_details["soft_skills"]
                    if isinstance(soft_data, dict) and "score" in soft_data:
                        soft_score = soft_data["score"]
            
            row = [
                str(resume.id),
                resume.hh_id or "",
                resume.name or "",
                resume.age or "",
                resume.city or "",
                resume.title or "",
                resume.salary or "",
                resume.currency or "",
                resume.preliminary_score or "",
                resume.ai_score or "",
                f"{resume.match_percentage:.1f}%" if resume.match_percentage else "",
                status,
                tags,
                resume.ai_summary or "",
                resume.match_explanation or "",
                "; ".join(resume.strengths) if resume.strengths else "",
                "; ".join(resume.weaknesses) if resume.weaknesses else "",
                resume.recommendation or "",
                tech_score,
                exp_score,
                edu_score,
                soft_score,
                "; ".join(resume.red_flags) if resume.red_flags else "",
                "; ".join(resume.ai_questions) if resume.ai_questions else "",
                "Yes" if resume.ai_generated_detected else "No",
                "Yes" if resume.analyzed else "No",
                resume.created_at.strftime("%Y-%m-%d %H:%M:%S") if resume.created_at else ""
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    logger.debug("Error calculating column width", error=str(e))
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("Excel export created", search_id=search_id, resumes_count=len(resumes))
        
        return output
    
    async def export_to_csv(self, search_id: str) -> BytesIO:
        """Export search results to CSV"""
        # Validate search_id format
        self._validate_search_id(search_id)
        
        # Get search
        search = await Search.get(search_id)
        if not search:
            raise NotFoundException("Search not found")
        
        # Get all resumes
        resumes = await Resume.find({"search_id": str(search.id)}).sort(-Resume.ai_score).to_list()
        
        # Create CSV
        output = BytesIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            "ID", "HH ID", "Name", "Age", "City", "Title", "Salary", "Currency",
            "Preliminary Score", "AI Score", "AI Summary", "AI Questions", "AI Generated",
            "Analyzed"
        ]
        writer.writerow(headers)
        
        # Add data
        for resume in resumes:
            row = [
                str(resume.id),
                resume.hh_id or "",
                resume.name or "",
                resume.age or "",
                resume.city or "",
                resume.title or "",
                resume.salary or "",
                resume.currency or "",
                resume.preliminary_score or "",
                resume.ai_score or "",
                resume.ai_summary or "",
                "; ".join(resume.ai_questions) if resume.ai_questions else "",
                "Yes" if resume.ai_generated_detected else "No",
                "Yes" if resume.analyzed else "No"
            ]
            writer.writerow(row)
        
        output.seek(0)
        
        logger.info("CSV export created", search_id=search_id, resumes_count=len(resumes))
        
        return output
    
    async def export_to_pdf(
        self,
        search_id: str,
        include_details: bool = True
    ) -> BytesIO:
        """Export search results to PDF report"""
        if not PDF_AVAILABLE:
            raise Exception("PDF export not available. Install reportlab: pip install reportlab")
        
        # Validate search_id format
        self._validate_search_id(search_id)
        
        # Get search
        search = await Search.get(search_id)
        if not search:
            raise NotFoundException("Search not found")
        
        # Get all resumes
        resumes = await Resume.find({"search_id": str(search.id)}).sort(-Resume.ai_score).to_list()
        
        # Create PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("Resume Analysis Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Search info
        info_style = styles['Normal']
        story.append(Paragraph(f"<b>Search Query:</b> {search.query}", info_style))
        story.append(Paragraph(f"<b>City:</b> {search.city}", info_style))
        story.append(Paragraph(f"<b>Total Found:</b> {len(resumes)}", info_style))
        story.append(Paragraph(f"<b>Date:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}", info_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Summary table
        if resumes:
            summary_data = [["Metric", "Value"]]
            analyzed_count = sum(1 for r in resumes if r.analyzed)
            avg_score = sum(r.ai_score for r in resumes if r.ai_score) / len([r for r in resumes if r.ai_score]) if any(r.ai_score for r in resumes) else 0
            
            summary_data.append(["Total Resumes", str(len(resumes))])
            summary_data.append(["Analyzed", str(analyzed_count)])
            summary_data.append(["Average AI Score", f"{avg_score:.2f}"])
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Resume details
        for idx, resume in enumerate(resumes[:50], 1):  # Limit to 50 for PDF
            if idx > 1:
                story.append(PageBreak())
            
            # Resume header
            header_style = ParagraphStyle(
                'ResumeHeader',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12
            )
            story.append(Paragraph(f"Resume #{idx}: {resume.name or 'Unknown'}", header_style))
            
            # Basic info
            basic_data = [
                ["Field", "Value"],
                ["Title", resume.title or "N/A"],
                ["Age", str(resume.age) if resume.age else "N/A"],
                ["City", resume.city or "N/A"],
                ["Salary", f"{resume.salary} {resume.currency}" if resume.salary else "N/A"],
                ["AI Score", str(resume.ai_score) if resume.ai_score else "N/A"],
                ["Match %", f"{resume.match_percentage:.1f}%" if resume.match_percentage else "N/A"]
            ]
            
            basic_table = Table(basic_data, colWidths=[2*inch, 4*inch])
            basic_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(basic_table)
            story.append(Spacer(1, 0.2*inch))
            
            if include_details:
                # AI Analysis
                if resume.match_explanation:
                    story.append(Paragraph("<b>Match Explanation:</b>", styles['Heading3']))
                    story.append(Paragraph(resume.match_explanation, styles['Normal']))
                    story.append(Spacer(1, 0.1*inch))
                
                if resume.strengths:
                    story.append(Paragraph("<b>Strengths:</b>", styles['Heading3']))
                    for strength in resume.strengths:
                        story.append(Paragraph(f"• {strength}", styles['Normal']))
                    story.append(Spacer(1, 0.1*inch))
                
                if resume.weaknesses:
                    story.append(Paragraph("<b>Weaknesses:</b>", styles['Heading3']))
                    for weakness in resume.weaknesses:
                        story.append(Paragraph(f"• {weakness}", styles['Normal']))
                    story.append(Spacer(1, 0.1*inch))
                
                if resume.recommendation:
                    story.append(Paragraph("<b>Recommendation:</b>", styles['Heading3']))
                    story.append(Paragraph(resume.recommendation, styles['Normal']))
                    story.append(Spacer(1, 0.1*inch))
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        logger.info("PDF export created", search_id=search_id, resumes_count=len(resumes))
        
        return output
    
    def _validate_vacancy_id(self, vacancy_id: str) -> None:
        """Validate vacancy_id format (ObjectId)"""
        from bson import ObjectId
        try:
            ObjectId(vacancy_id)
        except Exception:
            raise NotFoundException(f"Invalid vacancy ID format: {vacancy_id}")
    
    async def export_vacancy_report(
        self,
        vacancy_id: str,
        format: str = "excel"
    ) -> BytesIO:
        """Export vacancy report with all candidates"""
        from app.domain.entities.vacancy import Vacancy
        
        # Validate vacancy_id format
        self._validate_vacancy_id(vacancy_id)
        
        vacancy = await Vacancy.get(vacancy_id)
        if not vacancy:
            raise NotFoundException("Vacancy not found")
        
        if format == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = f"Vacancy: {vacancy.title[:30]}"
            
            # Vacancy info
            ws.append(["Vacancy Information"])
            ws.append(["Title", vacancy.title])
            ws.append(["Status", vacancy.status])
            ws.append(["City", vacancy.city])
            ws.append(["Created", vacancy.created_at.strftime("%Y-%m-%d")])
            ws.append([])
            
            # Headers
            headers = [
                "Name", "Title", "Age", "City", "AI Score", "Match %", "Status",
                "Match Explanation", "Recommendation", "Strengths", "Weaknesses"
            ]
            ws.append(headers)
            
            # Style headers
            for cell in ws[len(list(ws.iter_rows()))]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Add candidates
            for resume_id in vacancy.candidate_ids:
                resume = await Resume.get(resume_id)
                if resume:
                    candidate = await Candidate.find_one({"resume_id": resume_id})
                    status = candidate.status if candidate else "new"
                    
                    row = [
                        resume.name or "",
                        resume.title or "",
                        resume.age or "",
                        resume.city or "",
                        resume.ai_score or "",
                        f"{resume.match_percentage:.1f}%" if resume.match_percentage else "",
                        status,
                        resume.match_explanation or "",
                        resume.recommendation or "",
                        "; ".join(resume.strengths) if resume.strengths else "",
                        "; ".join(resume.weaknesses) if resume.weaknesses else ""
                    ]
                    ws.append(row)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
        else:
            raise ValueError(f"Unsupported format: {format}")


# Global service instance
export_service = ExportService()
